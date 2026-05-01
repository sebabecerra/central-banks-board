"""
Process the KOF ETH Zurich central-bank governors dataset into a clean flat file
with source URLs.

Source dataset
--------------
- `cbg_turnover_v23upload.xlsx`
- Sheet: `governors v2023`

Method
------
1. Read the KOF workbook and the target sheet.
2. Use row 0 as ISO3 codes and row 1 as country names.
3. Parse free-text governor entries from rows 2+.
4. Normalize ISO3 aliases and clean names.
5. Attach a source URL for each ISO3/entity.
6. Export the final dataset and a validation table of unresolved source URLs.

Outputs
-------
- `CM/data/kof_governors_with_sources.csv`
- `CM/data-aux/kof_missing_source_url.csv`
"""

import csv
import re
from pathlib import Path

import openpyxl
import pandas as pd

from kof_source_maps import SOURCE_URLS

BASE_DIR = Path(__file__).resolve().parents[1]
PROJECT_DIR = BASE_DIR.parents[0]
INPUT_FILE = PROJECT_DIR / "old" / "data" / "cbg_turnover_v23upload.xlsx"
SHEET_NAME = "governors v2023"
OUTPUT_DIR = BASE_DIR / "data"
OUTPUT_AUX_DIR = BASE_DIR / "data-aux"
OUTPUT_FILE = OUTPUT_DIR / "kof_governors_with_sources.csv"
UNRESOLVED_OUTPUT_FILE = OUTPUT_AUX_DIR / "kof_missing_source_url.csv"
BANK_LOOKUP_FILE = OUTPUT_AUX_DIR / "central_banks.csv"

STRICT_SOURCE_URL = False

ISO_ALIASES = {
    "AOA": "AGO",
    "ROM": "ROU",
    "BUR": "MMR",
    "ZAR": "COD",
    "SER": "SRB",
}


def normalize_iso(iso_value):
    iso = str(iso_value or "").strip().upper()
    return ISO_ALIASES.get(iso, iso)


def extract_year(value):
    if not value:
        return ""
    value = str(value).lower().strip()
    if any(token in value for token in ["present", "current", "ongoing"]):
        return "En cargo"
    match = re.search(r"\d{4}", value)
    return match.group() if match else ""


def clean_name(name):
    name = str(name or "")
    name = re.sub(
        r"^(dr\.|mr\.|mrs\.|ms\.|prof\.|ing\.|sir\s)\s*",
        "",
        name,
        flags=re.IGNORECASE,
    )
    name = re.sub(
        r"\s*\((?:[^)]*(?:reapp|reappointed|interim|acting|first|second|third|fourth|1st|2nd|3rd|4th|term|time)[^)]*)\)",
        "",
        name,
        flags=re.IGNORECASE,
    )
    name = re.sub(r"^\*+\s*", "", name)
    name = re.sub(r"\s+", " ", name).strip().rstrip(".,;")
    return name


def parse_cell(text, country_name, iso):
    if not text or str(text).strip() in ["None", ""]:
        return []

    text = str(text).strip()

    skip_keywords = [
        "established", "est. in", "est.in", "code:", "from 1985 onwards member",
        "member of beac", "(beac)", "(bceao)", "(eccb)", "no central bank",
        "code: -999", "exists since", "est. 1", "est.1",
    ]
    if any(pattern in text.lower() for pattern in skip_keywords):
        return []

    patterns = [
        r"(\d{1,2}-\d{1,2}-\d{4})\s+to\s+(\d{1,2}-\d{1,2}-\d{4})\s+(.+)",
        r"(\d{1,2}-\d{4})\s+to\s+(\d{1,2}-\d{4})\s+(.+)",
        r"(\d{1,2}-\d{4})\s+to\s+(present|current)\s+(.+)",
        r"(\d{4})\s+to\s+(\d{4})\s+(.+)",
        r"(\d{4})\s+to\s+(present|current)\s+(.+)",
        r"(\d{4})\s+-\s+(\d{4})\s+(.+)",
        r"(\d{4})\s+-\s+(present|current)\s+(.+)",
        r"(.+?)\s+(\d{1,2}-\d{4})\s+to\s+(\d{1,2}-\d{4})$",
        r"(.+?)\s+(\d{4})\s+to\s+(\d{4})$",
    ]

    results = []
    entries = re.split(r"\n", text)

    for entry in entries:
        entry = entry.strip()
        if not entry:
            continue

        for i, pattern in enumerate(patterns):
            match = re.search(pattern, entry, re.IGNORECASE)
            if not match:
                continue

            if i < 7:
                start_raw, end_raw, name = match.group(1), match.group(2), match.group(3)
            else:
                name, start_raw, end_raw = match.group(1), match.group(2), match.group(3)

            name = clean_name(name)
            if len(name) < 2:
                continue

            start_year = extract_year(start_raw)
            end_year = extract_year(end_raw)
            if not start_year:
                continue

            results.append(
                {
                    "iso3": iso,
                    "country": country_name,
                    "name": name,
                    "position": "President / Governor",
                    "start_year": start_year,
                    "end_year": end_year,
                    "status": "Actual" if end_year == "En cargo" else "Histórico",
                    "source_url": SOURCE_URLS.get(iso, ""),
                }
            )
            break

    return results


def build_country_map(rows):
    iso_codes = rows[0]
    country_names = rows[1]

    countries = {}
    for i, (iso, name) in enumerate(zip(iso_codes, country_names)):
        if iso and name:
            countries[i] = {
                "iso": normalize_iso(iso),
                "name": str(name).strip(),
            }
    return countries


def build_unresolved_source_df(countries):
    country_map_df = pd.DataFrame(countries.values()).drop_duplicates().sort_values(["iso", "name"])
    country_map_df["source_url"] = country_map_df["iso"].map(SOURCE_URLS).fillna("")
    return country_map_df[country_map_df["source_url"] == ""].copy()


def build_country_to_bank_lookup():
    if not BANK_LOOKUP_FILE.exists():
        return {}

    lookup_df = pd.read_csv(BANK_LOOKUP_FILE, sep=";")
    required = {"country", "central_bank"}
    if not required.issubset(lookup_df.columns):
        return {}

    lookup_df = lookup_df[list(required)].copy()
    lookup_df["country"] = lookup_df["country"].astype(str).str.strip()
    lookup_df["central_bank"] = lookup_df["central_bank"].astype(str).str.strip()
    lookup_df = lookup_df[
        (lookup_df["country"] != "") &
        (lookup_df["country"].str.lower() != "nan") &
        (lookup_df["central_bank"] != "") &
        (lookup_df["central_bank"].str.lower() != "nan")
    ].drop_duplicates(subset=["country"], keep="first")

    return dict(zip(lookup_df["country"], lookup_df["central_bank"]))


def parse_all_records(rows, countries):
    all_records = []
    for row in rows[2:]:
        for col_idx, value in enumerate(row):
            if col_idx not in countries:
                continue
            if not value or str(value).strip() in ["None", ""]:
                continue
            country_info = countries[col_idx]
            all_records.extend(parse_cell(value, country_info["name"], country_info["iso"]))
    return all_records


def deduplicate_records(records):
    seen = set()
    unique = []
    for record in records:
        key = (
            record["iso3"],
            record["name"].lower(),
            record["start_year"],
            record["end_year"],
        )
        if key not in seen:
            seen.add(key)
            unique.append(record)
    return unique


def print_preview(name, df, rows=10):
    print(f"{name}: {df.shape}")
    if df.empty:
        print(f"{name} is empty\n")
        return
    print(df.head(rows).to_string(index=False))
    print()


def build_final_output(processed_df, country_to_bank):
    final_columns = [
        "country",
        "central_bank_name",
        "name",
        "position",
        "start_year",
        "end_year",
        "source_dataset",
        "source_method",
        "source_page",
        "source_detail",
    ]

    if processed_df.empty:
        return pd.DataFrame(columns=final_columns)

    final_df = processed_df.copy()
    final_df["central_bank_name"] = final_df["country"].map(country_to_bank).fillna("")
    final_df["source_dataset"] = "kof"
    final_df["source_method"] = "kof_workbook_governors_v2023"
    final_df["source_page"] = final_df["source_url"].fillna("")
    final_df["source_detail"] = final_df["iso3"].fillna("")

    final_df = final_df[
        final_columns
    ].drop_duplicates().sort_values(
        ["country", "central_bank_name", "start_year", "name"],
        na_position="last",
    ).reset_index(drop=True)

    return final_df


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {INPUT_FILE}")

    workbook = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"La hoja {SHEET_NAME!r} no existe en el archivo")

    worksheet = workbook[SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))

    countries = build_country_map(rows)
    unresolved_source_df = build_unresolved_source_df(countries)

    if STRICT_SOURCE_URL and not unresolved_source_df.empty:
        raise ValueError(
            "Existen ISO3 sin source_url. Revisa el archivo de pendientes o completa SOURCE_URLS."
        )

    all_records = parse_all_records(rows, countries)
    unique = deduplicate_records(all_records)

    processed_df = pd.DataFrame(unique)
    processed_df = processed_df.sort_values(["country", "start_year", "name"], na_position="last")
    country_to_bank = build_country_to_bank_lookup()
    final_df = build_final_output(processed_df, country_to_bank)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_AUX_DIR.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "country", "central_bank_name", "name", "position", "start_year",
        "end_year", "source_dataset", "source_method", "source_page",
        "source_detail",
    ]
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(final_df.to_dict(orient="records"))

    unresolved_source_df.to_csv(UNRESOLVED_OUTPUT_FILE, index=False, sep=";")

    saved_df = pd.read_csv(OUTPUT_FILE, sep=";")
    print_preview("kof_governors_with_sources", saved_df, rows=20)

    print("Resumen general")
    print("-" * 40)
    print("Total registros        :", len(saved_df))
    print("Países únicos          :", saved_df["country"].nunique())
    print("Bancos con match       :", saved_df["central_bank_name"].fillna("").astype(str).str.strip().ne("").sum())
    print("Gobernadores actuales  :", (saved_df["end_year"] == "En cargo").sum())
    print("Gobernadores históricos:", (saved_df["end_year"] != "En cargo").sum())


if __name__ == "__main__":
    main()
